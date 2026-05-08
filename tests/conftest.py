import os
import pytest
from datetime import date
from openpyxl import Workbook


@pytest.fixture
def sample_workbook():
    """
    Minimal workbook matching the confirmed grid layout:
      Row 1: CW labels  |  Row 2: Weekday names
      Col C: Projekt Migration (filter)  |  Col D: Team Mitglied (name)
      Col F (idx 6) = 2026-04-27 (Mon, CW18) ... Col L (idx 12) = 2026-05-05 (Tue)

    Data rows (5 members, 3 marked):
      Row 3  Alice   marked  absent F(Apr27), G(Apr28)
      Row 4  Bob     marked  absent H(Apr29), I(Apr30), J(May1)
      Row 5  Carol   marked  no absences
      Row 6  Dave    NOT marked
      Row 7  Eve     NOT marked (col C = "y")
      Row 8  "Alice " (trailing space)  marked  absent K(May4)  → merged with row 3
      Row 9  ""      marked, empty name → skipped row warning
    """
    wb = Workbook()
    ws = wb.active

    # Row 1: CW labels
    for col, label in zip(range(6, 13), ["KW18", "KW18", "KW18", "KW18", "KW18", "KW19", "KW19"]):
        ws.cell(row=1, column=col, value=label)

    # Row 2: Weekday names (Mon=6, Tue=7, Wed=8, Thu=9, Fri=10, Mon=11, Tue=12)
    for col, day in zip(range(6, 13), ["Mo", "Di", "Mi", "Do", "Fr", "Mo", "Di"]):
        ws.cell(row=2, column=col, value=day)

    # Row 3: Alice – marked, absent Apr27 + Apr28
    ws.cell(row=3, column=3, value="x")
    ws.cell(row=3, column=4, value="Alice")
    ws.cell(row=3, column=6, value="x")   # Apr 27
    ws.cell(row=3, column=7, value="x")   # Apr 28

    # Row 4: Bob – marked, absent Apr29 + Apr30 + May1 (contiguous)
    ws.cell(row=4, column=3, value="x")
    ws.cell(row=4, column=4, value="Bob")
    ws.cell(row=4, column=8, value="x")   # Apr 29
    ws.cell(row=4, column=9, value="x")   # Apr 30
    ws.cell(row=4, column=10, value="x")  # May 1

    # Row 5: Carol – marked, no absences
    ws.cell(row=5, column=3, value="x")
    ws.cell(row=5, column=4, value="Carol")

    # Row 6: Dave – NOT marked (empty col C) – excluded
    ws.cell(row=6, column=4, value="Dave")
    ws.cell(row=6, column=6, value="x")

    # Row 7: Eve – NOT marked (col C = "y") – excluded
    ws.cell(row=7, column=3, value="y")
    ws.cell(row=7, column=4, value="Eve")
    ws.cell(row=7, column=6, value="x")

    # Row 8: "Alice " (trailing space) – same person as Alice, absent May4
    ws.cell(row=8, column=3, value="x")
    ws.cell(row=8, column=4, value="Alice ")
    ws.cell(row=8, column=11, value="x")  # May 4

    # Row 9: empty name – marked but name is blank → skipped row
    ws.cell(row=9, column=3, value="x")
    ws.cell(row=9, column=4, value="")

    return wb


@pytest.fixture
def sample_xlsx(sample_workbook, tmp_path):
    path = tmp_path / "test_absences.xlsx"
    sample_workbook.save(str(path))
    return str(path)


@pytest.fixture
def app(sample_xlsx, tmp_path):
    state_path = str(tmp_path / "state.json")
    from absence_dashboard.app import create_app
    flask_app = create_app(sample_xlsx, state_path=state_path)
    flask_app.config["TESTING"] = True
    return flask_app


@pytest.fixture
def client(app):
    return app.test_client()
