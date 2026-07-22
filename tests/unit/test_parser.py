"""
TDD: Tests for absence_dashboard/parser.py
Write BEFORE implementation; confirm they FAIL before writing parser.py.
"""
import pytest
from datetime import date
from openpyxl import Workbook

from absence_dashboard.parser import build_date_map, parse_members, PersonAbsence, SkippedRow


# ---------------------------------------------------------------------------
# build_date_map
# ---------------------------------------------------------------------------

class TestBuildDateMap:
    def _make_ws(self, columns):
        """columns: list of (kw_label, weekday_abbrev) tuples, one per column starting at F(6)."""
        wb = Workbook()
        ws = wb.active
        for offset, (kw, wd) in enumerate(columns):
            col = 6 + offset
            ws.cell(row=1, column=col, value=kw)
            ws.cell(row=2, column=col, value=wd)
        return ws

    def test_col_f_is_base_date(self):
        ws = self._make_ws([("KW18", "Mo")])
        dm = build_date_map(ws, year=2026)
        assert dm[6] == date(2026, 4, 27)

    def test_col_g_is_next_working_day(self):
        ws = self._make_ws([("KW18", "Mo"), ("KW18", "Di")])
        dm = build_date_map(ws, year=2026)
        assert dm[7] == date(2026, 4, 28)  # Tuesday

    def test_col_k_skips_weekend(self):
        # Col 6=Mon Apr27, 7=Tue Apr28, 8=Wed Apr29, 9=Thu Apr30,
        # 10=Fri May1, 11=Mon May4 (skip Sat May2 + Sun May3)
        ws = self._make_ws([
            ("KW18", "Mo"), ("KW18", "Di"), ("KW18", "Mi"), ("KW18", "Do"), ("KW18", "Fr"),
            ("KW19", "Mo"),
        ])
        dm = build_date_map(ws, year=2026)
        assert dm[10] == date(2026, 5, 1)   # Friday
        assert dm[11] == date(2026, 5, 4)   # Monday (weekend skipped)

    def test_no_columns_below_6(self):
        ws = self._make_ws([])
        dm = build_date_map(ws, year=2026)
        assert 5 not in dm
        assert 6 not in dm

    def test_returns_all_columns_from_6(self):
        ws = self._make_ws([("KW18", "Mo"), ("KW18", "Di"), ("KW18", "Mi")])
        dm = build_date_map(ws, year=2026)
        assert set(dm.keys()) == {6, 7, 8}

    def test_uses_actual_header_dates_not_hardcoded_sequence(self):
        # Regression test for a real production bug: a column's date MUST be derived
        # from that column's own header cells (Row 1 CW label, Row 2 weekday), never
        # from a fixed starting date plus a sequential working-day increment — the real
        # file's actual date range can (and did) differ from any such hardcoded
        # assumption, silently mis-dating every absence in the sheet.
        ws = self._make_ws([("KW30", "Mo")])
        dm = build_date_map(ws, year=2026)
        assert dm[6] == date(2026, 7, 20)
        assert dm[6] != date(2026, 4, 27)

    def test_year_rolls_over_when_week_number_decreases(self):
        ws = self._make_ws([("KW52", "Mo"), ("KW1", "Mo")])
        dm = build_date_map(ws, year=2026)
        assert dm[6] == date(2026, 12, 21)
        assert dm[7] == date(2027, 1, 4)

    def test_column_with_unrecognized_header_is_skipped(self):
        ws = self._make_ws([("KW18", "Mo"), ("", ""), ("KW18", "Mi")])
        dm = build_date_map(ws, year=2026)
        assert set(dm.keys()) == {6, 8}

    def test_full_german_weekday_names_recognized(self):
        # Regression test for a real production file: Row 2 can hold full German day
        # names ("Montag", "Dienstag", ...) instead of the two-letter abbreviations
        # ("Mo", "Di", ...) the original format assumed — both must resolve correctly.
        ws = self._make_ws([
            ("KW18", "Montag"), ("18", "Dienstag"), ("18", "Mittwoch"),
            ("18", "Donnerstag"), ("18", "Freitag"),
        ])
        dm = build_date_map(ws, year=2026)
        assert dm[6] == date(2026, 4, 27)
        assert dm[7] == date(2026, 4, 28)
        assert dm[8] == date(2026, 4, 29)
        assert dm[9] == date(2026, 4, 30)
        assert dm[10] == date(2026, 5, 1)

    def test_bare_integer_week_number_recognized(self):
        # Regression test: Row 1 can hold a bare integer (18) instead of a "KW18"-style
        # string — both must resolve to the same week.
        ws = self._make_ws([(18, "Mo")])
        dm = build_date_map(ws, year=2026)
        assert dm[6] == date(2026, 4, 27)

    def test_merged_week_number_cell_forward_filled(self):
        # Regression test for a real production bug: Excel commonly merges the week-
        # number cell across all 5 weekday columns (visually shown once, spanning
        # Mon-Fri). A merged cell's value only physically exists in its leftmost column
        # — every other cell it covers reads as blank. Row 1 must forward-fill the most
        # recently seen week number across those blank cells, or every day but the first
        # of each week silently drops out of the map (only Monday ever got dated).
        ws = self._make_ws([
            ("KW18", "Montag"), (None, "Dienstag"), (None, "Mittwoch"),
            (None, "Donnerstag"), (None, "Freitag"), ("KW19", "Montag"),
        ])
        dm = build_date_map(ws, year=2026)
        assert dm[6] == date(2026, 4, 27)
        assert dm[7] == date(2026, 4, 28)
        assert dm[8] == date(2026, 4, 29)
        assert dm[9] == date(2026, 4, 30)
        assert dm[10] == date(2026, 5, 1)
        assert dm[11] == date(2026, 5, 4)

    def test_merged_week_number_cell_across_year_rollover(self):
        ws = self._make_ws([
            ("KW52", "Montag"), (None, "Dienstag"), ("KW1", "Montag"), (None, "Dienstag"),
        ])
        dm = build_date_map(ws, year=2026)
        assert dm[6] == date(2026, 12, 21)
        assert dm[7] == date(2026, 12, 22)
        assert dm[8] == date(2027, 1, 4)
        assert dm[9] == date(2027, 1, 5)


# ---------------------------------------------------------------------------
# parse_members
# ---------------------------------------------------------------------------

class TestParseMembers:
    def test_all_named_rows_included(self, sample_workbook):
        ws = sample_workbook.active
        members, _ = parse_members(ws)
        names = {m.name for m in members}
        assert "Alice" in names
        assert "Bob" in names
        assert "Carol" in names
        assert "Dave" in names
        assert "Eve" in names

    def test_exactly_five_members(self, sample_workbook):
        ws = sample_workbook.active
        members, _ = parse_members(ws)
        assert len(members) == 5

    def test_x_detection_case_insensitive(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=6, value="KW18")
        ws.cell(row=2, column=6, value="Mo")
        ws.cell(row=3, column=3, value="X")   # uppercase X in filter
        ws.cell(row=3, column=4, value="Frank")
        ws.cell(row=3, column=6, value="X")   # uppercase absence marker
        members, _ = parse_members(ws)
        assert len(members) == 1
        assert members[0].name == "Frank"
        assert date(2026, 4, 27) in members[0].absence_days

    def test_name_whitespace_stripped(self, sample_workbook):
        ws = sample_workbook.active
        members, _ = parse_members(ws)
        names = {m.name for m in members}
        # "Alice " (row 8) should be stripped and merged with "Alice" (row 3)
        assert "Alice" in names
        assert "Alice " not in names

    def test_multi_row_same_name_aggregated(self, sample_workbook):
        ws = sample_workbook.active
        members, _ = parse_members(ws)
        alice = next(m for m in members if m.name == "Alice")
        # Row 3: Apr27, Apr28; Row 8: May4 (after name strip, same person)
        assert date(2026, 4, 27) in alice.absence_days
        assert date(2026, 4, 28) in alice.absence_days
        assert date(2026, 5, 4) in alice.absence_days

    def test_empty_name_row_skipped(self, sample_workbook):
        ws = sample_workbook.active
        _, skipped = parse_members(ws)
        assert len(skipped) == 1
        assert skipped[0].row == 9

    def test_rows_1_and_2_skipped(self):
        wb = Workbook()
        ws = wb.active
        # Put "x" in col C row 1 and row 2 — should be ignored
        ws.cell(row=1, column=3, value="x")
        ws.cell(row=1, column=4, value="HeaderPerson")
        ws.cell(row=2, column=3, value="x")
        ws.cell(row=2, column=4, value="HeaderPerson2")
        members, _ = parse_members(ws)
        assert len(members) == 0

    def test_bob_has_correct_absence_days(self, sample_workbook):
        ws = sample_workbook.active
        members, _ = parse_members(ws)
        bob = next(m for m in members if m.name == "Bob")
        assert date(2026, 4, 29) in bob.absence_days
        assert date(2026, 4, 30) in bob.absence_days
        assert date(2026, 5, 1) in bob.absence_days
        assert len(bob.absence_days) == 3

    def test_carol_has_no_absence_days(self, sample_workbook):
        ws = sample_workbook.active
        members, _ = parse_members(ws)
        carol = next(m for m in members if m.name == "Carol")
        assert carol.absence_days == []

    def test_returns_person_absence_and_skipped_row_types(self, sample_workbook):
        ws = sample_workbook.active
        members, skipped = parse_members(ws)
        assert all(isinstance(m, PersonAbsence) for m in members)
        assert all(isinstance(s, SkippedRow) for s in skipped)

    def test_is_migration_member_field_present(self, sample_workbook):
        ws = sample_workbook.active
        members, _ = parse_members(ws)
        for m in members:
            assert hasattr(m, "is_migration_member"), f"{m.name} missing is_migration_member"

    def test_migration_members_flagged_true(self, sample_workbook):
        ws = sample_workbook.active
        members, _ = parse_members(ws)
        migration_names = {"Alice", "Bob", "Carol"}
        for m in members:
            if m.name in migration_names:
                assert m.is_migration_member is True, f"{m.name} should be migration member"

    def test_non_migration_members_flagged_false(self, sample_workbook):
        ws = sample_workbook.active
        members, _ = parse_members(ws)
        non_migration_names = {"Dave", "Eve"}
        for m in members:
            if m.name in non_migration_names:
                assert m.is_migration_member is False, f"{m.name} should NOT be migration member"

    def test_same_name_any_row_marked_gives_true(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=6, value="KW18")
        ws.cell(row=2, column=6, value="Mo")
        # Row 3: Grace NOT marked
        ws.cell(row=3, column=4, value="Grace")
        ws.cell(row=3, column=6, value="x")
        # Row 4: Grace MARKED
        ws.cell(row=4, column=3, value="x")
        ws.cell(row=4, column=4, value="Grace")
        members, _ = parse_members(ws)
        grace = next(m for m in members if m.name == "Grace")
        assert grace.is_migration_member is True

    def test_uses_real_header_dates_for_a_late_starting_sheet(self):
        # Regression test for the real production bug: a sheet whose data columns start
        # at a week far from the old hardcoded BASE_DATE (April 27, 2026) must still be
        # dated correctly, derived from its own header cells.
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=6, value="KW30")
        ws.cell(row=2, column=6, value="Mo")
        ws.cell(row=3, column=3, value="x")
        ws.cell(row=3, column=4, value="Hilde")
        ws.cell(row=3, column=6, value="x")
        members, _ = parse_members(ws)
        hilde = next(m for m in members if m.name == "Hilde")
        assert date(2026, 7, 20) in hilde.absence_days
