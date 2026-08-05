"""
Integration tests for absence_dashboard/app.py Flask routes.
TDD: Written BEFORE implementation; confirmed failing before app.py is complete.
"""
import json
import sys
import pytest
from datetime import date
from unittest.mock import MagicMock


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def get_member(data, name):
    return next((m for m in data["members"] if m["name"] == name), None)


# ---------------------------------------------------------------------------
# Fixtures for cumul risk / sole coverage tests (T020 / US2)
# Dates are computed relative to today so the fixture stays valid regardless
# of when the suite runs (calendar_weeks only spans today..CW53 of this year).
# 3 consecutive weeks starting from the week after today: Alice absent
# week1 Mon - week2 Mon (8 cal days, critical); Bob absent week2 Mon -
# week3 Mon (8 cal days, critical). They share week2's Monday. Carol never
# absent.
# ---------------------------------------------------------------------------

def _monday_of_iso_week(year, week):
    from datetime import date
    return date.fromisocalendar(year, week, 1)


@pytest.fixture
def critical_weeks():
    from datetime import date, timedelta
    today_iso = date.today().isocalendar()
    start_week = today_iso.week + 1
    year = today_iso.year
    mondays = [
        _monday_of_iso_week(year, start_week),
        _monday_of_iso_week(year, start_week + 1),
        _monday_of_iso_week(year, start_week + 2),
    ]
    week_numbers = [start_week, start_week + 1, start_week + 2]
    weekday_columns = [
        [mon + timedelta(days=offset) for offset in range(5)] for mon in mondays
    ]
    return {"year": year, "week_numbers": week_numbers, "days": weekday_columns}


@pytest.fixture
def critical_workbook(critical_weeks):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active

    all_days = [d for week in critical_weeks["days"] for d in week]
    cw_labels = [f"KW{wn}" for wn in critical_weeks["week_numbers"] for _ in range(5)]
    for col, label in zip(range(6, 21), cw_labels):
        ws.cell(row=1, column=col, value=label)

    weekdays = ["Mo", "Di", "Mi", "Do", "Fr"] * 3
    for col, day in zip(range(6, 21), weekdays):
        ws.cell(row=2, column=col, value=day)

    # Row 3: Alice – absent week1 Mon-Fri + week2 Mon (cols 6-11)
    ws.cell(row=3, column=3, value="x")
    ws.cell(row=3, column=4, value="Alice")
    for col in range(6, 12):
        ws.cell(row=3, column=col, value="x")

    # Row 4: Bob – absent week2 Mon-Fri + week3 Mon (cols 11-16)
    ws.cell(row=4, column=3, value="x")
    ws.cell(row=4, column=4, value="Bob")
    for col in range(11, 17):
        ws.cell(row=4, column=col, value="x")

    # Row 5: Carol – marked, no absences
    ws.cell(row=5, column=3, value="x")
    ws.cell(row=5, column=4, value="Carol")

    return wb


@pytest.fixture
def critical_xlsx(critical_workbook, tmp_path):
    path = tmp_path / "critical_absences.xlsx"
    critical_workbook.save(str(path))
    return str(path)


@pytest.fixture
def critical_app(critical_xlsx, tmp_path):
    state_path = str(tmp_path / "critical_state.json")
    from absence_dashboard.app import create_app
    flask_app = create_app(critical_xlsx, state_path=state_path)
    flask_app.config["TESTING"] = True
    return flask_app


@pytest.fixture
def critical_client(critical_app):
    return critical_app.test_client()


# ---------------------------------------------------------------------------
# GET /api/dashboard  (T014 / US1)
# ---------------------------------------------------------------------------

class TestGetDashboard:
    def test_status_200(self, client):
        rv = client.get("/api/dashboard")
        assert rv.status_code == 200

    def test_exactly_five_members(self, client):
        data = client.get("/api/dashboard").get_json()
        assert len(data["members"]) == 5

    def test_member_names_correct(self, client):
        data = client.get("/api/dashboard").get_json()
        names = {m["name"] for m in data["members"]}
        assert names == {"Alice", "Bob", "Carol", "Dave", "Eve"}

    def test_alice_merged_blocks(self, client):
        data = client.get("/api/dashboard").get_json()
        alice = get_member(data, "Alice")
        assert alice is not None
        # Row 3: Apr27-28; Row 8 (same person): May4 → two non-overlapping blocks
        starts = {b["start"] for b in alice["merged_blocks"]}
        assert "2026-04-27" in starts
        assert "2026-05-04" in starts

    def test_alice_first_block_end(self, client):
        data = client.get("/api/dashboard").get_json()
        alice = get_member(data, "Alice")
        block_27 = next(b for b in alice["merged_blocks"] if b["start"] == "2026-04-27")
        assert block_27["end"] == "2026-04-28"

    def test_bob_single_merged_block(self, client):
        data = client.get("/api/dashboard").get_json()
        bob = get_member(data, "Bob")
        assert len(bob["merged_blocks"]) == 1
        assert bob["merged_blocks"][0]["start"] == "2026-04-29"
        assert bob["merged_blocks"][0]["end"] == "2026-05-01"

    def test_carol_no_merged_blocks(self, client):
        data = client.get("/api/dashboard").get_json()
        carol = get_member(data, "Carol")
        assert carol["merged_blocks"] == []

    def test_merged_blocks_non_overlapping(self, client):
        data = client.get("/api/dashboard").get_json()
        for member in data["members"]:
            blocks = sorted(member["merged_blocks"], key=lambda b: b["start"])
            for i in range(len(blocks) - 1):
                assert blocks[i]["end"] < blocks[i + 1]["start"]

    def test_calendar_weeks_starts_at_current_week(self, client):
        data = client.get("/api/dashboard").get_json()
        weeks = data["calendar_weeks"]
        assert len(weeks) > 0
        today = date.today()
        iso = today.isocalendar()
        assert weeks[0]["year"] == iso.year
        assert weeks[0]["week_number"] == iso.week

    def test_calendar_weeks_ends_at_cw53_2026(self, client):
        data = client.get("/api/dashboard").get_json()
        weeks = data["calendar_weeks"]
        assert weeks[-1]["year"] == 2026
        assert weeks[-1]["week_number"] == 53

    def test_no_duplicate_calendar_weeks(self, client):
        data = client.get("/api/dashboard").get_json()
        keys = [(w["year"], w["week_number"]) for w in data["calendar_weeks"]]
        assert len(keys) == len(set(keys))

    def test_calendar_weeks_have_days_field(self, client):
        data = client.get("/api/dashboard").get_json()
        for week in data["calendar_weeks"]:
            assert "days" in week, f"CW{week['week_number']} missing 'days' field"
            assert len(week["days"]) == 5, f"CW{week['week_number']} should have 5 days"

    def test_calendar_week_days_are_mon_to_fri(self, client):
        from datetime import date, datetime
        data = client.get("/api/dashboard").get_json()
        first_week = data["calendar_weeks"][0]
        days = first_week["days"]
        for i, day_str in enumerate(days):
            d = date.fromisoformat(day_str)
            assert d.weekday() == i, f"Day {i} should be weekday {i}, got {d.weekday()}"

    def test_calendar_week_days_match_start(self, client):
        data = client.get("/api/dashboard").get_json()
        for week in data["calendar_weeks"]:
            assert week["days"][0] == week["start"], (
                f"First day should equal start for {week['label']}"
            )

    def test_calendar_week_label_format(self, client):
        import re
        data = client.get("/api/dashboard").get_json()
        pattern = re.compile(r"^CW\d{1,2} \| \d{1,2} [A-Z][a-z]{2}$")
        for week in data["calendar_weeks"]:
            assert pattern.match(week["label"]), (
                f"Label '{week['label']}' does not match 'CW[N] | D Mon' format"
            )

    def test_skipped_rows_present(self, client):
        data = client.get("/api/dashboard").get_json()
        assert "skipped_rows" in data
        assert len(data["skipped_rows"]) == 1

    def test_member_fields_present(self, client):
        data = client.get("/api/dashboard").get_json()
        for m in data["members"]:
            assert "name" in m
            assert "merged_blocks" in m
            assert "clusters" in m

    def test_is_migration_member_field_present(self, client):
        data = client.get("/api/dashboard").get_json()
        for m in data["members"]:
            assert "is_migration_member" in m, f"{m['name']} missing is_migration_member"

    def test_migration_members_have_true_flag(self, client):
        data = client.get("/api/dashboard").get_json()
        migration_names = {"Alice", "Bob", "Carol"}
        for m in data["members"]:
            if m["name"] in migration_names:
                assert m["is_migration_member"] is True

    def test_non_migration_members_have_false_flag(self, client):
        data = client.get("/api/dashboard").get_json()
        non_migration_names = {"Dave", "Eve"}
        for m in data["members"]:
            if m["name"] in non_migration_names:
                assert m["is_migration_member"] is False

    def test_all_five_members_returned(self, client):
        data = client.get("/api/dashboard").get_json()
        names = {m["name"] for m in data["members"]}
        assert "Dave" in names
        assert "Eve" in names


# ---------------------------------------------------------------------------
# Cluster endpoints  (T029 / US4)
# ---------------------------------------------------------------------------

class TestPostClusters:
    def test_create_cluster_returns_201(self, client):
        rv = client.post("/api/clusters",
                         json={"name": "Backend", "members": ["Alice", "Bob"]})
        assert rv.status_code == 201

    def test_cluster_appears_in_dashboard(self, client):
        client.post("/api/clusters", json={"name": "Backend", "members": ["Alice"]})
        data = client.get("/api/dashboard").get_json()
        alice = get_member(data, "Alice")
        assert "Backend" in alice["clusters"]

    def test_duplicate_cluster_name_returns_400(self, client):
        client.post("/api/clusters", json={"name": "Backend", "members": ["Alice"]})
        rv = client.post("/api/clusters", json={"name": "Backend", "members": ["Bob"]})
        assert rv.status_code == 400

    def test_unknown_member_returns_400(self, client):
        rv = client.post("/api/clusters",
                         json={"name": "Backend", "members": ["Unknown"]})
        assert rv.status_code == 400


class TestPutClusters:
    def test_update_cluster_members(self, client):
        client.post("/api/clusters", json={"name": "Backend", "members": ["Alice"]})
        rv = client.put("/api/clusters/Backend", json={"members": ["Alice", "Bob"]})
        assert rv.status_code == 200
        data = client.get("/api/dashboard").get_json()
        bob = get_member(data, "Bob")
        assert "Backend" in bob["clusters"]

    def test_update_unknown_cluster_returns_404(self, client):
        rv = client.put("/api/clusters/Nonexistent", json={"members": ["Alice"]})
        assert rv.status_code == 404

    def test_update_with_unknown_member_returns_400(self, client):
        client.post("/api/clusters", json={"name": "Backend", "members": ["Alice"]})
        rv = client.put("/api/clusters/Backend", json={"members": ["Unknown"]})
        assert rv.status_code == 400


class TestDeleteClusters:
    def test_delete_existing_cluster_returns_200(self, client):
        client.post("/api/clusters", json={"name": "Backend", "members": ["Alice"]})
        rv = client.delete("/api/clusters/Backend")
        assert rv.status_code == 200

    def test_delete_removes_cluster_from_dashboard(self, client):
        client.post("/api/clusters", json={"name": "Backend", "members": ["Alice"]})
        client.delete("/api/clusters/Backend")
        data = client.get("/api/dashboard").get_json()
        alice = get_member(data, "Alice")
        assert "Backend" not in alice["clusters"]

    def test_delete_unknown_cluster_returns_404(self, client):
        rv = client.delete("/api/clusters/Nonexistent")
        assert rv.status_code == 404


class TestMemberInMultipleClusters:
    def test_member_in_two_clusters_appears_in_both(self, client):
        client.post("/api/clusters", json={"name": "Backend", "members": ["Alice"]})
        client.post("/api/clusters", json={"name": "DevOps", "members": ["Alice"]})
        data = client.get("/api/dashboard").get_json()
        alice = get_member(data, "Alice")
        assert "Backend" in alice["clusters"]
        assert "DevOps" in alice["clusters"]


# ---------------------------------------------------------------------------
# GET /api/clusters  (T038 / Polish)
# ---------------------------------------------------------------------------

class TestGetEndpoints:
    def test_get_clusters_returns_200(self, client):
        rv = client.get("/api/clusters")
        assert rv.status_code == 200
        assert "clusters" in rv.get_json()


# ---------------------------------------------------------------------------
# Cumul group endpoints  (T012 / US1)
# ---------------------------------------------------------------------------

class TestCumulGroupsCRUD:
    def test_get_cumul_groups_returns_200(self, client):
        rv = client.get("/api/cumul-groups")
        assert rv.status_code == 200
        assert rv.get_json()["cumul_groups"] == []

    def test_create_cumul_group_returns_201(self, client):
        rv = client.post(
            "/api/cumul-groups",
            json={"name": "Backend Coverage", "members": ["Alice", "Bob"]},
        )
        assert rv.status_code == 201
        assert rv.get_json()["cumul_groups"][0]["name"] == "Backend Coverage"

    def test_created_group_appears_in_get(self, client):
        client.post(
            "/api/cumul-groups",
            json={"name": "Backend Coverage", "members": ["Alice", "Bob"]},
        )
        data = client.get("/api/cumul-groups").get_json()
        assert data["cumul_groups"][0]["members"] == ["Alice", "Bob"]

    def test_create_with_fewer_than_two_members_returns_400(self, client):
        rv = client.post(
            "/api/cumul-groups", json={"name": "Solo", "members": ["Alice"]},
        )
        assert rv.status_code == 400

    def test_create_with_empty_name_returns_400(self, client):
        rv = client.post(
            "/api/cumul-groups", json={"name": "", "members": ["Alice", "Bob"]},
        )
        assert rv.status_code == 400

    def test_create_with_unknown_member_returns_400(self, client):
        rv = client.post(
            "/api/cumul-groups",
            json={"name": "Backend Coverage", "members": ["Alice", "Unknown"]},
        )
        assert rv.status_code == 400

    def test_create_duplicate_name_returns_409(self, client):
        client.post(
            "/api/cumul-groups",
            json={"name": "Backend Coverage", "members": ["Alice", "Bob"]},
        )
        rv = client.post(
            "/api/cumul-groups",
            json={"name": "Backend Coverage", "members": ["Carol", "Dave"]},
        )
        assert rv.status_code == 409

    def test_create_duplicate_member_set_returns_409(self, client):
        client.post(
            "/api/cumul-groups",
            json={"name": "Backend Coverage", "members": ["Alice", "Bob"]},
        )
        rv = client.post(
            "/api/cumul-groups",
            json={"name": "Backend Coverage v2", "members": ["Bob", "Alice"]},
        )
        assert rv.status_code == 409

    def test_create_active_from_without_active_to_returns_400(self, client):
        rv = client.post(
            "/api/cumul-groups",
            json={
                "name": "Timeboxed", "members": ["Alice", "Bob"],
                "active_from": "2026-06-01",
            },
        )
        assert rv.status_code == 400

    def test_update_cumul_group_rename(self, client):
        client.post(
            "/api/cumul-groups",
            json={"name": "Backend Coverage", "members": ["Alice", "Bob"]},
        )
        rv = client.put(
            "/api/cumul-groups/Backend Coverage",
            json={"name": "Backend Coverage v2"},
        )
        assert rv.status_code == 200
        assert rv.get_json()["cumul_groups"][0]["name"] == "Backend Coverage v2"

    def test_update_cumul_group_members(self, client):
        client.post(
            "/api/cumul-groups",
            json={"name": "Backend Coverage", "members": ["Alice", "Bob"]},
        )
        rv = client.put(
            "/api/cumul-groups/Backend Coverage",
            json={"members": ["Carol", "Dave"]},
        )
        assert rv.status_code == 200
        assert rv.get_json()["cumul_groups"][0]["members"] == ["Carol", "Dave"]

    def test_update_unknown_group_returns_404(self, client):
        rv = client.put(
            "/api/cumul-groups/Nonexistent", json={"name": "Renamed"},
        )
        assert rv.status_code == 404

    def test_update_with_unknown_member_returns_400(self, client):
        client.post(
            "/api/cumul-groups",
            json={"name": "Backend Coverage", "members": ["Alice", "Bob"]},
        )
        rv = client.put(
            "/api/cumul-groups/Backend Coverage",
            json={"members": ["Alice", "Unknown"]},
        )
        assert rv.status_code == 400

    def test_update_rename_to_existing_name_returns_409(self, client):
        client.post(
            "/api/cumul-groups",
            json={"name": "Backend Coverage", "members": ["Alice", "Bob"]},
        )
        client.post(
            "/api/cumul-groups",
            json={"name": "Frontend Coverage", "members": ["Carol", "Dave"]},
        )
        rv = client.put(
            "/api/cumul-groups/Backend Coverage",
            json={"name": "Frontend Coverage"},
        )
        assert rv.status_code == 409

    def test_delete_cumul_group_returns_200(self, client):
        client.post(
            "/api/cumul-groups",
            json={"name": "Backend Coverage", "members": ["Alice", "Bob"]},
        )
        rv = client.delete("/api/cumul-groups/Backend Coverage")
        assert rv.status_code == 200
        assert rv.get_json()["cumul_groups"] == []

    def test_delete_unknown_group_returns_404(self, client):
        rv = client.delete("/api/cumul-groups/Nonexistent")
        assert rv.status_code == 404

    def test_cumul_groups_appear_in_dashboard(self, client):
        client.post(
            "/api/cumul-groups",
            json={"name": "Backend Coverage", "members": ["Alice", "Bob"]},
        )
        data = client.get("/api/dashboard").get_json()
        assert data["cumul_groups"][0]["name"] == "Backend Coverage"


# ---------------------------------------------------------------------------
# Cumul risk weeks visibility  (T020 / US2)
# ---------------------------------------------------------------------------

class TestCumulRiskWeeks:
    def test_dashboard_member_has_cumul_risks_field(self, critical_client):
        data = critical_client.get("/api/dashboard").get_json()
        for m in data["members"]:
            assert "cumul_risks" in m

    def test_shared_critical_week_flagged(self, critical_client, critical_weeks):
        critical_client.post(
            "/api/cumul-groups",
            json={"name": "Risk Group", "members": ["Alice", "Bob"]},
        )
        data = critical_client.get("/api/dashboard").get_json()
        alice = get_member(data, "Alice")
        bob = get_member(data, "Bob")
        week2 = critical_weeks["week_numbers"][1]
        assert {"group": "Risk Group", "week_number": week2} in alice["cumul_risks"]
        assert {"group": "Risk Group", "week_number": week2} in bob["cumul_risks"]

    def test_non_overlapping_weeks_not_flagged(self, critical_client, critical_weeks):
        critical_client.post(
            "/api/cumul-groups",
            json={"name": "Risk Group", "members": ["Alice", "Bob"]},
        )
        data = critical_client.get("/api/dashboard").get_json()
        alice = get_member(data, "Alice")
        weeks = {r["week_number"] for r in alice["cumul_risks"] if r["group"] == "Risk Group"}
        week1, _, week3 = critical_weeks["week_numbers"]
        assert week1 not in weeks
        assert week3 not in weeks

    def test_member_not_in_group_has_no_risk_for_that_group(self, critical_client):
        critical_client.post(
            "/api/cumul-groups",
            json={"name": "Risk Group", "members": ["Alice", "Bob"]},
        )
        data = critical_client.get("/api/dashboard").get_json()
        carol = get_member(data, "Carol")
        assert carol["cumul_risks"] == []

    def test_no_groups_means_no_risks(self, critical_client):
        data = critical_client.get("/api/dashboard").get_json()
        for m in data["members"]:
            assert m["cumul_risks"] == []


# ---------------------------------------------------------------------------
# Sole coverage visibility  (T020 / US2)
# ---------------------------------------------------------------------------

class TestSoleCoverage:
    def test_dashboard_member_has_sole_coverage_field(self, critical_client):
        data = critical_client.get("/api/dashboard").get_json()
        for m in data["members"]:
            assert "sole_coverage" in m

    def test_sole_present_member_flagged(self, critical_client, critical_weeks):
        critical_client.post(
            "/api/cumul-groups",
            json={"name": "Sole Group", "members": ["Alice", "Bob", "Carol"]},
        )
        data = critical_client.get("/api/dashboard").get_json()
        carol = get_member(data, "Carol")
        week2 = critical_weeks["week_numbers"][1]
        assert {"group": "Sole Group", "week_number": week2} in carol["sole_coverage"]

    def test_weeks_where_carol_not_sole_are_not_flagged(self, critical_client, critical_weeks):
        critical_client.post(
            "/api/cumul-groups",
            json={"name": "Sole Group", "members": ["Alice", "Bob", "Carol"]},
        )
        data = critical_client.get("/api/dashboard").get_json()
        carol = get_member(data, "Carol")
        weeks = {r["week_number"] for r in carol["sole_coverage"] if r["group"] == "Sole Group"}
        week1, _, week3 = critical_weeks["week_numbers"]
        assert week1 not in weeks
        assert week3 not in weeks

    def test_two_member_group_flags_both_directions(self, critical_client, critical_weeks):
        critical_client.post(
            "/api/cumul-groups",
            json={"name": "Risk Group", "members": ["Alice", "Bob"]},
        )
        data = critical_client.get("/api/dashboard").get_json()
        alice = get_member(data, "Alice")
        bob = get_member(data, "Bob")
        week1, _, week3 = critical_weeks["week_numbers"]
        assert {"group": "Risk Group", "week_number": week1} in bob["sole_coverage"]
        assert {"group": "Risk Group", "week_number": week3} in alice["sole_coverage"]


# ---------------------------------------------------------------------------
# POST /api/refresh  (T034 / US5)
# ---------------------------------------------------------------------------

def _write_workbook_with_members(path, member_names):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=6, value="KW18")
    ws.cell(row=2, column=6, value="Mo")
    for i, name in enumerate(member_names):
        row = 3 + i
        ws.cell(row=row, column=3, value="x")
        ws.cell(row=row, column=4, value=name)
    wb.save(path)


class TestRefresh:
    def test_refresh_returns_200(self, client):
        rv = client.post("/api/refresh")
        assert rv.status_code == 200

    def test_refresh_preserves_clusters(self, client):
        client.post("/api/clusters", json={"name": "Backend", "members": ["Alice"]})
        client.post("/api/refresh")
        data = client.get("/api/dashboard").get_json()
        alice = get_member(data, "Alice")
        assert "Backend" in alice["clusters"]


class TestRefreshCumulGroupCleanup:
    def test_removed_member_dropped_and_reported(self, client, app):
        client.post("/api/cumul-groups", json={"name": "G1", "members": ["Alice", "Bob", "Carol"]})
        _write_workbook_with_members(app.config["EXCEL_SOURCE"], ["Alice", "Bob"])
        rv = client.post("/api/refresh")
        data = rv.get_json()
        removed = data["removed_stale_references"]
        assert {"type": "cumul_group_member", "entry": {"group": "G1", "member": "Carol"}} in removed
        group = next(g for g in data["cumul_groups"] if g["name"] == "G1")
        assert group["members"] == ["Alice", "Bob"]

    def test_group_removed_when_below_two_valid_members(self, client, app):
        client.post("/api/cumul-groups", json={"name": "G2", "members": ["Alice", "Bob"]})
        _write_workbook_with_members(app.config["EXCEL_SOURCE"], ["Alice"])
        rv = client.post("/api/refresh")
        data = rv.get_json()
        removed = data["removed_stale_references"]
        assert any(
            r["type"] == "cumul_group" and r["entry"]["name"] == "G2"
            for r in removed
        )
        assert not any(g["name"] == "G2" for g in data["cumul_groups"])

    def test_unaffected_group_unchanged(self, client, app):
        client.post("/api/cumul-groups", json={"name": "G3", "members": ["Alice", "Bob"]})
        _write_workbook_with_members(app.config["EXCEL_SOURCE"], ["Alice", "Bob", "Carol"])
        rv = client.post("/api/refresh")
        data = rv.get_json()
        group = next(g for g in data["cumul_groups"] if g["name"] == "G3")
        assert group["members"] == ["Alice", "Bob"]
        removed = data["removed_stale_references"]
        assert not any(r["entry"].get("group") == "G3" for r in removed)


# ---------------------------------------------------------------------------
# Phase endpoints  (T047 / US6)
# ---------------------------------------------------------------------------

class TestGetPhases:
    def test_get_phases_returns_200(self, client):
        rv = client.get("/api/phases")
        assert rv.status_code == 200

    def test_get_phases_has_phases_key(self, client):
        rv = client.get("/api/phases")
        assert "phases" in rv.get_json()

    def test_initial_phases_empty(self, client):
        rv = client.get("/api/phases")
        assert rv.get_json()["phases"] == []


class TestPostPhases:
    def test_add_valid_phase_returns_201(self, client):
        rv = client.post("/api/phases",
                         json={"name": "Go-Live", "start_date": "2026-06-22", "end_date": "2026-06-26"})
        assert rv.status_code == 201

    def test_add_valid_phase_appears_in_list(self, client):
        client.post("/api/phases",
                    json={"name": "Go-Live", "start_date": "2026-06-22", "end_date": "2026-06-26"})
        rv = client.get("/api/phases")
        phases = rv.get_json()["phases"]
        assert any(p["name"] == "Go-Live" for p in phases)

    def test_add_phase_duplicate_name_returns_400(self, client):
        client.post("/api/phases",
                    json={"name": "Go-Live", "start_date": "2026-06-22", "end_date": "2026-06-26"})
        rv = client.post("/api/phases",
                         json={"name": "Go-Live", "start_date": "2026-07-01", "end_date": "2026-07-05"})
        assert rv.status_code == 400

    def test_add_phase_end_before_start_returns_400(self, client):
        rv = client.post("/api/phases",
                         json={"name": "BadPhase", "start_date": "2026-07-10", "end_date": "2026-07-05"})
        assert rv.status_code == 400

    def test_add_phase_single_day_returns_201(self, client):
        rv = client.post("/api/phases",
                         json={"name": "Kickoff", "start_date": "2026-06-01", "end_date": "2026-06-01"})
        assert rv.status_code == 201

    def test_add_phase_overlapping_allowed(self, client):
        client.post("/api/phases",
                    json={"name": "Sprint 10", "start_date": "2026-06-15", "end_date": "2026-06-26"})
        rv = client.post("/api/phases",
                         json={"name": "Go-Live", "start_date": "2026-06-22", "end_date": "2026-06-26"})
        assert rv.status_code == 201


class TestDeletePhase:
    def test_delete_existing_phase_returns_200(self, client):
        client.post("/api/phases",
                    json={"name": "Go-Live", "start_date": "2026-06-22", "end_date": "2026-06-26"})
        rv = client.delete("/api/phases/Go-Live")
        assert rv.status_code == 200

    def test_delete_removes_phase(self, client):
        client.post("/api/phases",
                    json={"name": "Go-Live", "start_date": "2026-06-22", "end_date": "2026-06-26"})
        client.delete("/api/phases/Go-Live")
        rv = client.get("/api/phases")
        phases = rv.get_json()["phases"]
        assert not any(p["name"] == "Go-Live" for p in phases)

    def test_delete_nonexistent_phase_returns_404(self, client):
        rv = client.delete("/api/phases/Nonexistent")
        assert rv.status_code == 404

    def test_delete_url_encoded_name(self, client):
        client.post("/api/phases",
                    json={"name": "Go Live", "start_date": "2026-06-22", "end_date": "2026-06-26"})
        import urllib.parse
        encoded = urllib.parse.quote("Go Live")
        rv = client.delete(f"/api/phases/{encoded}")
        assert rv.status_code == 200


class TestDashboardIncludesPhases:
    def test_dashboard_has_phases_key(self, client):
        data = client.get("/api/dashboard").get_json()
        assert "phases" in data

    def test_dashboard_phases_initially_empty(self, client):
        data = client.get("/api/dashboard").get_json()
        assert data["phases"] == []

    def test_dashboard_reflects_added_phase(self, client):
        client.post("/api/phases",
                    json={"name": "Go-Live", "start_date": "2026-06-22", "end_date": "2026-06-26"})
        data = client.get("/api/dashboard").get_json()
        assert any(p["name"] == "Go-Live" for p in data["phases"])


class TestRefreshPreservesPhases:
    def test_refresh_preserves_phases(self, client):
        client.post("/api/phases",
                    json={"name": "Go-Live", "start_date": "2026-06-22", "end_date": "2026-06-26"})
        client.post("/api/refresh")
        data = client.get("/api/dashboard").get_json()
        assert any(p["name"] == "Go-Live" for p in data["phases"])


# ---------------------------------------------------------------------------
# PUT /api/clusters/<name> with rename  (T057 / Phase 11)
# ---------------------------------------------------------------------------

class TestPutClustersWithRename:
    def test_rename_only_returns_200(self, client):
        client.post("/api/clusters", json={"name": "Backend", "members": ["Alice"]})
        rv = client.put("/api/clusters/Backend", json={"name": "Core"})
        assert rv.status_code == 200

    def test_rename_updates_cluster_name(self, client):
        client.post("/api/clusters", json={"name": "Backend", "members": ["Alice"]})
        client.put("/api/clusters/Backend", json={"name": "Core"})
        data = client.get("/api/clusters").get_json()
        names = [c["name"] for c in data["clusters"]]
        assert "Core" in names
        assert "Backend" not in names

    def test_rename_preserves_members(self, client):
        client.post("/api/clusters", json={"name": "Backend", "members": ["Alice", "Bob"]})
        client.put("/api/clusters/Backend", json={"name": "Core"})
        data = client.get("/api/clusters").get_json()
        core = next(c for c in data["clusters"] if c["name"] == "Core")
        assert set(core["members"]) == {"Alice", "Bob"}

    def test_update_members_only_returns_200(self, client):
        client.post("/api/clusters", json={"name": "Backend", "members": ["Alice"]})
        rv = client.put("/api/clusters/Backend", json={"members": ["Alice", "Carol"]})
        assert rv.status_code == 200

    def test_update_members_only_preserves_name(self, client):
        client.post("/api/clusters", json={"name": "Backend", "members": ["Alice"]})
        client.put("/api/clusters/Backend", json={"members": ["Alice", "Carol"]})
        data = client.get("/api/clusters").get_json()
        assert any(c["name"] == "Backend" for c in data["clusters"])

    def test_rename_and_update_members(self, client):
        client.post("/api/clusters", json={"name": "Backend", "members": ["Alice"]})
        rv = client.put("/api/clusters/Backend", json={"name": "Core", "members": ["Carol"]})
        assert rv.status_code == 200
        data = client.get("/api/clusters").get_json()
        core = next(c for c in data["clusters"] if c["name"] == "Core")
        assert core["members"] == ["Carol"]

    def test_rename_to_duplicate_name_returns_400(self, client):
        client.post("/api/clusters", json={"name": "Backend", "members": ["Alice"]})
        client.post("/api/clusters", json={"name": "Frontend", "members": ["Bob"]})
        rv = client.put("/api/clusters/Backend", json={"name": "Frontend"})
        assert rv.status_code == 400

    def test_rename_to_same_name_is_ok(self, client):
        client.post("/api/clusters", json={"name": "Backend", "members": ["Alice"]})
        rv = client.put("/api/clusters/Backend", json={"name": "Backend"})
        assert rv.status_code == 200

    def test_unknown_cluster_returns_404(self, client):
        rv = client.put("/api/clusters/Nonexistent", json={"name": "X"})
        assert rv.status_code == 404

    def test_invalid_member_returns_400(self, client):
        client.post("/api/clusters", json={"name": "Backend", "members": ["Alice"]})
        rv = client.put("/api/clusters/Backend", json={"members": ["Unknown"]})
        assert rv.status_code == 400

    def test_dashboard_reflects_rename(self, client):
        client.post("/api/clusters", json={"name": "Backend", "members": ["Alice"]})
        client.put("/api/clusters/Backend", json={"name": "Core"})
        data = client.get("/api/dashboard").get_json()
        alice = get_member(data, "Alice")
        assert "Core" in alice["clusters"]
        assert "Backend" not in alice["clusters"]


# ---------------------------------------------------------------------------
# PUT /api/phases/<name>  (T058 / Phase 11)
# ---------------------------------------------------------------------------

class TestPutPhases:
    def test_update_name_only_returns_200(self, client):
        client.post("/api/phases",
                    json={"name": "Go-Live", "start_date": "2026-06-22", "end_date": "2026-06-26"})
        rv = client.put("/api/phases/Go-Live", json={"name": "Launch"})
        assert rv.status_code == 200

    def test_update_name_only_renames_phase(self, client):
        client.post("/api/phases",
                    json={"name": "Go-Live", "start_date": "2026-06-22", "end_date": "2026-06-26"})
        client.put("/api/phases/Go-Live", json={"name": "Launch"})
        phases = client.get("/api/phases").get_json()["phases"]
        names = [p["name"] for p in phases]
        assert "Launch" in names
        assert "Go-Live" not in names

    def test_update_dates_only_returns_200(self, client):
        client.post("/api/phases",
                    json={"name": "Go-Live", "start_date": "2026-06-22", "end_date": "2026-06-26"})
        rv = client.put("/api/phases/Go-Live",
                        json={"start_date": "2026-06-23", "end_date": "2026-06-27"})
        assert rv.status_code == 200

    def test_update_dates_only_changes_dates(self, client):
        client.post("/api/phases",
                    json={"name": "Go-Live", "start_date": "2026-06-22", "end_date": "2026-06-26"})
        client.put("/api/phases/Go-Live",
                   json={"start_date": "2026-06-23", "end_date": "2026-06-27"})
        phases = client.get("/api/phases").get_json()["phases"]
        go_live = next(p for p in phases if p["name"] == "Go-Live")
        assert go_live["start_date"] == "2026-06-23"
        assert go_live["end_date"] == "2026-06-27"

    def test_update_all_fields(self, client):
        client.post("/api/phases",
                    json={"name": "Go-Live", "start_date": "2026-06-22", "end_date": "2026-06-26"})
        rv = client.put("/api/phases/Go-Live",
                        json={"name": "Launch", "start_date": "2026-07-01",
                              "end_date": "2026-07-05"})
        assert rv.status_code == 200
        phases = client.get("/api/phases").get_json()["phases"]
        launch = next(p for p in phases if p["name"] == "Launch")
        assert launch["start_date"] == "2026-07-01"
        assert launch["end_date"] == "2026-07-05"

    def test_duplicate_name_returns_400(self, client):
        client.post("/api/phases",
                    json={"name": "Go-Live", "start_date": "2026-06-22", "end_date": "2026-06-26"})
        client.post("/api/phases",
                    json={"name": "Sprint 10", "start_date": "2026-06-15", "end_date": "2026-06-21"})
        rv = client.put("/api/phases/Go-Live", json={"name": "Sprint 10"})
        assert rv.status_code == 400

    def test_end_before_start_returns_400(self, client):
        client.post("/api/phases",
                    json={"name": "Go-Live", "start_date": "2026-06-22", "end_date": "2026-06-26"})
        rv = client.put("/api/phases/Go-Live",
                        json={"start_date": "2026-07-10", "end_date": "2026-07-05"})
        assert rv.status_code == 400

    def test_unknown_phase_returns_404(self, client):
        rv = client.put("/api/phases/Nonexistent", json={"name": "X"})
        assert rv.status_code == 404

    def test_state_persisted_after_update(self, app, client):
        client.post("/api/phases",
                    json={"name": "Go-Live", "start_date": "2026-06-22", "end_date": "2026-06-26"})
        client.put("/api/phases/Go-Live", json={"name": "Launch"})
        state = app.config["STATE"]
        assert any(p["name"] == "Launch" for p in state.phases)
        assert not any(p["name"] == "Go-Live" for p in state.phases)

    def test_dashboard_reflects_phase_update(self, client):
        client.post("/api/phases",
                    json={"name": "Go-Live", "start_date": "2026-06-22", "end_date": "2026-06-26"})
        client.put("/api/phases/Go-Live", json={"name": "Launch"})
        data = client.get("/api/dashboard").get_json()
        assert any(p["name"] == "Launch" for p in data["phases"])
        assert not any(p["name"] == "Go-Live" for p in data["phases"])


# ---------------------------------------------------------------------------
# last_loaded timestamp  (T068 / FR-025)
# ---------------------------------------------------------------------------

class TestLastLoaded:
    def test_dashboard_has_last_loaded(self, client):
        data = client.get("/api/dashboard").get_json()
        assert "last_loaded" in data

    def test_last_loaded_is_iso_datetime_string(self, client):
        from datetime import datetime
        data = client.get("/api/dashboard").get_json()
        ll = data["last_loaded"]
        assert isinstance(ll, str)
        dt = datetime.strptime(ll, "%Y-%m-%dT%H:%M:%S")
        assert dt is not None

    def test_refresh_includes_last_loaded(self, client):
        rv = client.post("/api/refresh")
        assert rv.status_code == 200
        data = rv.get_json()
        assert "last_loaded" in data

    def test_refresh_last_loaded_is_iso_datetime_string(self, client):
        from datetime import datetime
        rv = client.post("/api/refresh")
        ll = rv.get_json()["last_loaded"]
        dt = datetime.strptime(ll, "%Y-%m-%dT%H:%M:%S")
        assert dt is not None

    def test_refresh_last_loaded_not_earlier_than_initial(self, client):
        data1 = client.get("/api/dashboard").get_json()
        ll1 = data1["last_loaded"]
        data2 = client.post("/api/refresh").get_json()
        ll2 = data2["last_loaded"]
        assert ll2 >= ll1


# ---------------------------------------------------------------------------
# resolve_launch_source() — CLI arg vs launch_config.json precedence (feature 002)
# ---------------------------------------------------------------------------

class TestResolveLaunchSource:
    def test_cli_arg_takes_precedence_over_launch_config(self, tmp_path):
        cli_xlsx = tmp_path / "cli.xlsx"
        cli_xlsx.write_text("dummy")
        config_path = tmp_path / "launch_config.json"
        config_path.write_text(json.dumps({"excel_source": "config.xlsx", "port": 9999}))

        from absence_dashboard.app import resolve_launch_source
        source, port = resolve_launch_source(str(cli_xlsx), None, config_path=str(config_path))

        assert source == str(cli_xlsx)
        assert port == 5002  # existing CLI default, NOT launch_config's 9999

    def test_no_cli_arg_falls_back_to_launch_config(self, tmp_path):
        xlsx = tmp_path / "absences.xlsx"
        xlsx.write_text("dummy")
        config_path = tmp_path / "launch_config.json"
        config_path.write_text(json.dumps({"excel_source": str(xlsx), "port": 6100}))

        from absence_dashboard.app import resolve_launch_source
        source, port = resolve_launch_source(None, None, config_path=str(config_path))

        assert source == str(xlsx)
        assert port == 6100

    def test_explicit_cli_port_overrides_launch_config_port(self, tmp_path):
        xlsx = tmp_path / "absences.xlsx"
        xlsx.write_text("dummy")
        config_path = tmp_path / "launch_config.json"
        config_path.write_text(json.dumps({"excel_source": str(xlsx), "port": 6100}))

        from absence_dashboard.app import resolve_launch_source
        source, port = resolve_launch_source(None, 7000, config_path=str(config_path))

        assert port == 7000

    def test_no_cli_arg_and_missing_launch_config_raises(self, tmp_path):
        config_path = tmp_path / "missing.json"

        from absence_dashboard.app import resolve_launch_source
        with pytest.raises(FileNotFoundError):
            resolve_launch_source(None, None, config_path=str(config_path))

    def test_nonexistent_cli_path_raises_file_not_found(self, tmp_path):
        config_path = tmp_path / "launch_config.json"
        config_path.write_text(json.dumps({"excel_source": "irrelevant.xlsx"}))

        from absence_dashboard.app import resolve_launch_source
        with pytest.raises(FileNotFoundError):
            resolve_launch_source(str(tmp_path / "does_not_exist.xlsx"), None, config_path=str(config_path))
